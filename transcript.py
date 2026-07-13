import re, sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
# how to run(example): python3 transcript.py workshop3.vtt
# each transcript becomes a tab in coding.xlsx (makes the file if missing)
def load_codes():
    f = Path(__file__).parent / "codes.txt"
    if not f.exists():
        sys.exit("Missing codes.txt (put it next to transcript.py)")
    return [line.strip() for line in f.read_text(encoding="utf-8").splitlines() if line.strip()]
code_list = load_codes()

def split_long(text, maxsent=5):
    # split into sentences; if more than maxsent, break into chunks of maxsent
    import re as _re
    parts = _re.split(r'(?<=[.!?])\s+', text.strip())
    parts = [p for p in parts if p]
    if len(parts) <= maxsent:
        return [text]
    chunks = []
    for i in range(0, len(parts), maxsent):
        chunks.append(" ".join(parts[i:i+maxsent]))
    return chunks

timestamp = re.compile(r'^\d{1,2}:\d{2}:\d{2}\.\d')
speaker = re.compile(r'^([^:]+?)\s*:\s*(.*)', re.S)
def read_transcript(path):
    text = Path(path).read_text(encoding="utf-8").replace("\r\n", "\n").replace("\r", "\n")
    turns = []
    for block in text.split("\n\n"):
        block = block.strip()
        if not block:
            continue
        time = ""
        said = []
        for line in block.split("\n"):
            if timestamp.match(line.strip()):
                time = line.strip()
            else:
                said.append(line)
        text_said = " ".join(said).strip()
        if not text_said:
            continue
        m = speaker.match(text_said)
        if m:
            person, quote = m.group(1).strip(), m.group(2).strip()
        else:
            person, quote = "", text_said
        if person.lower().startswith(("researcher","interviewer")):
            person = "Interviewer"
        turns.append((person, time, quote))
    return turns
# get the transcript name from the command line
if len(sys.argv) < 2:
    sys.exit("Usage: python3 transcript.py yourfile.vtt")
input_file = sys.argv[1]
# this transcript becomes one tab, named after the file
tab_name = Path(input_file).stem[:31]
# all transcripts go in one file
output_file = str(Path(input_file).with_name("coding.xlsx"))
turns = read_transcript(input_file)
if not turns:
    sys.exit("Nothing parsed - check the transcript format.")
bold = Font(bold=True, color="FFFFFF", name="Arial")
blue = PatternFill("solid", fgColor="2F5496")
wrap = Alignment(wrap_text=True, vertical="top")
top = Alignment(vertical="top")
# open the file if it exists, else start a new one
if Path(output_file).exists():
    book = load_workbook(output_file)
else:
    book = Workbook()
    book.remove(book.active)
#codes sheet: Code, Description, Times used
if "Codes" not in book.sheetnames:
    cs = book.create_sheet("Codes", 0)
    cs.cell(row=1, column=1, value="Code").font = bold
    cs.cell(row=1, column=1).fill = blue
    cs.cell(row=1, column=2, value="Description").font = bold
    cs.cell(row=1, column=2).fill = blue
    cs.cell(row=1, column=3, value="Times used").font = bold
    cs.cell(row=1, column=3).fill = blue
    cs.column_dimensions["A"].width = 40
    cs.column_dimensions["B"].width = 55
    cs.column_dimensions["C"].width = 12
    for i, code in enumerate(code_list, start=2):
        cs.cell(row=i, column=1, value=code).alignment = wrap
#one sheet per transcript, 3 cols each coder + 4 combined
if tab_name in book.sheetnames:
    sys.exit(f"A tab named '{tab_name}' already exists. Delete or rename it first.")
sheet = book.create_sheet(tab_name)
hdr = ["Speaker", "Timestamp", "Utterance", "Dr.Ming 1", "Dr.Ming 2", "Dr.Ming 3", "Omir 1", "Omir 2", "Omir 3", "Hawi 1", "Hawi 2", "Hawi 3", "Combined 1", "Combined 2", "Combined 3", "Combined 4"]
for j, h in enumerate(hdr, start=1):
    c = sheet.cell(row=1, column=j, value=h)
    c.font = bold
    c.fill = blue
sheet.column_dimensions["A"].width = 12
sheet.column_dimensions["B"].width = 11
sheet.column_dimensions["C"].width = 70
for j in range(4, 17):
    sheet.column_dimensions[get_column_letter(j)].width = 26
sheet.freeze_panes = "D2"
row = 2
# collect PIDs for the participant list
pids = []
for person, time, quote in turns:
    if person and person != "Interviewer" and person not in pids:
        pids.append(person)
for person, time, quote in turns:
    # split long PARTICIPANT turns (>5 sentences) into separate rows so theyre
    # easier to code. dont split the Interviewer.
    if person != "Interviewer":
        chunks = split_long(quote, 5)
    else:
        chunks = [quote]
    for chunk in chunks:
        sheet.cell(row=row, column=1, value=person).alignment = top
        sheet.cell(row=row, column=2, value=time).alignment = top
        sheet.cell(row=row, column=3, value=chunk).alignment = wrap
        row += 1
# participant list in far-right header cell (R1), doesnt shift any data
sheet.cell(row=1, column=18, value="Participants: " + ", ".join(pids))
sheet.cell(row=1, column=18).font = Font(bold=True, color="2F5496")
sheet.column_dimensions["R"].width = 40
# yellow highlight any code cell not in the Codes list
yellow = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
rule = FormulaRule(formula=['AND(D2<>"",COUNTIF(Codes!$A$2:$A$200,D2)=0)'], fill=yellow)
sheet.conditional_formatting.add(f"D2:P{row-1}", rule)
# pull from code sheet (all code cols D-P)
drop = DataValidation(type="list", formula1="=Codes!$A$2:$A$200", allow_blank=True)
sheet.add_data_validation(drop)
drop.add(f"D2:P{row-1}")
#update the count to cover every transcript tab
tabs = [s for s in book.sheetnames if s not in ("Codes", "GetCodes", "NewCodes")]
cs = book["Codes"]
for i in range(2, len(code_list) + 2):
    count = "+".join(f"COUNTIF('{t}'!$D$2:$P$1000, A{i})" for t in tabs)
    cs.cell(row=i, column=3, value=f"={count}")
# get codes: auto-lists every applied code once you paste in one formula.
# to use get teh formual from getcodes_formula.txt
# copy all, then in Excel click getcode in  A2 paste
if "GetCodes" in book.sheetnames:
    book.remove(book["GetCodes"])
gc = book.create_sheet("GetCodes")
gc.append(["Code", "Quote", "Speaker", "Timestamp"])
for c in gc[1]:
    c.font = bold
    c.fill = blue
gc.column_dimensions["A"].width = 40
gc.column_dimensions["B"].width = 70
gc.column_dimensions["C"].width = 12
gc.column_dimensions["D"].width = 12
gc["A2"] = "Paste the formula from getcodes_formula.txt into this cell (A2)."
gc["A2"].font = Font(italic=True, color="808080")
# stack every code from each tab (D-P), then keep only rows that have a code (no blanks)
pieces = []
for t in tabs:
    for col in ["M", "N", "O", "P"]:
        pieces.append(f"HSTACK('{t}'!{col}2:{col}1000,'{t}'!C2:C1000,'{t}'!A2:A1000,'{t}'!B2:B1000)")
stacked = "VSTACK(" + ",".join(pieces) + ")"
formula = f'=LET(all,{stacked},FILTER(all,INDEX(all,,1)<>""))'
(Path(output_file).parent / "getcodes_formula.txt").write_text(formula, encoding="utf-8")
# new codes: lists any code typed in Combined (M-P) thats not in the Codes list yet.
# paste the formula from newcodes_formula.txt into NewCodes cell A2.
if "NewCodes" in book.sheetnames:
    book.remove(book["NewCodes"])
nc = book.create_sheet("NewCodes")
nc.append(["New code spotted (not in Codes list yet)"])
nc["A1"].font = bold
nc["A1"].fill = blue
nc.column_dimensions["A"].width = 50
nc["A2"] = "Paste the formula from newcodes_formula.txt into this cell (A2)."
nc["A2"].font = Font(italic=True, color="808080")
comb = ",".join(f"'{t}'!{c}2:{c}1000" for t in tabs for c in ["M", "N", "O", "P"])
new_formula = f'=LET(all,VSTACK({comb}),filled,FILTER(all,all<>""),uniq,UNIQUE(filled),FILTER(uniq,ISNA(MATCH(uniq,Codes!$A$2:$A$200,0))))'
(Path(output_file).parent / "newcodes_formula.txt").write_text(new_formula, encoding="utf-8")
book.save(output_file)
print(f"Added tab '{tab_name}' to {output_file} ({row-2} turns)")
print(f"Tabs now: {', '.join(tabs)}")
missing = sum(1 for person, time, quote in turns if not person)
if missing:
    print(f"WARNING: {missing} turns had no speaker (check the format)")